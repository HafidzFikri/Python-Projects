import openpyxl

#### Jalankan untuk membuat file excel ##############
###### Making new workbook
wb = openpyxl.Workbook()
#Making new sheet for the data

data = wb.create_sheet("Data")
data["A1"] = "No"
data["B1"] = "Date"
data["C1"] = "Mie Ayam"
data["D1"] = "Bakso"
data["E1"] = "Es Teh"
data["F1"] = "Air Mineral"
data["G1"] = "Total"

data["I1"] = "Harga"
data["I2"] = "Mie Ayam"
data["J2"] = 12000
data["I3"] = "Bakso"
data["J3"] = 15000
data["I4"] = "Es Teh"
data["J4"] = 5000
data["I5"] = "Air Mineral"
data["J5"] = 2000

#wb.save("data.xlsx")
#wb.close()
#############################################

wk = openpyxl.load_workbook("Data Kasir.xlsx")
ws = wk['Data']

no = 1
tanggal = 8
bulan = 8
tahun = 2023
mieayam = [10,5,0,12,7,8,13,4,0,5]
bakso = [5,3,12,6,7,0,0,12,9,10]
esteh = [9,8,7,6,9,10,5,6,8,10]
air = [2,3,4,2,6,2,4,7,8,4]

j2 = ws['J2'].value
H_mieayam = j2.replace("Rp. ","")
j3 = ws['J3'].value
H_bakso = j3.replace("Rp. ","")
j4 = ws['J4'].value
H_esteh = j4.replace("Rp. ","")
j5 = ws['J5'].value
H_air = j5.replace("Rp. ","")
listharga = [H_mieayam,H_bakso,H_esteh,H_air]


for row_idx, (mieayam,bakso,esteh,air) in enumerate(zip(mieayam,bakso,esteh,air), start=2):
    if tanggal > 31:
        tanggal = 1
        bulan = bulan + 1
    total = f"Rp. {int(mieayam)*int(H_mieayam) + int(bakso)*int(H_bakso) + int(esteh)*int(H_esteh) + int(air)*int(H_air)}"
    ws.cell(row=row_idx, column=1, value=no)
    ws.cell(row=row_idx, column=2, value=f"{tanggal}/{bulan}/2023")
    ws.cell(row=row_idx, column=3, value=mieayam)
    ws.cell(row=row_idx, column=4, value=bakso)
    ws.cell(row=row_idx, column=5, value=esteh)
    ws.cell(row=row_idx, column=6, value=air)
    ws.cell(row=row_idx, column=7, value=total)
    no += 1
    tanggal += 1


while True:
    data_jual_mieayam = data_jual_bakso = data_jual_esteh = data_jual_air = None
    pesanan = {"Mie Ayam":data_jual_mieayam, "Bakso":data_jual_bakso, 
            "Es Teh": data_jual_esteh, "Air Mineral": data_jual_air}
    listpesanan = {}
    while True:
        for pesan,jumlah in pesanan.items():
            while True:
                jumlah = input(f"Masukkan jumlah pesanan {pesan.upper()} : ")
                if not jumlah:
                    jumlah = 0
                    break
                else:
                    try:
                        jumlah = int(jumlah)
                        break
                    except ValueError:
                        print("Masukkan Angka")
            listpesanan[pesan] = jumlah    
        break 

    total = f"Rp. {int(listpesanan['Mie Ayam'])*int(H_mieayam) + int(listpesanan['Bakso'])*int(H_bakso) + int(listpesanan['Es Teh'])*int(H_esteh) + int(listpesanan['Air Mineral'])*int(H_air)}" 
    row_jual = 11    
    while ws.cell(row=row_jual, column=1).value is not None:
        row_jual += 1
    no = row_jual - 1

    column = 3
    tanggal += 1
    if tanggal > 31:
        bulan = bulan + 1
        tanggal = 1
    if bulan > 12:
        tahun = tahun + 1
        bulan = 1
    tanggal = tanggal

    ws.cell(row=row_jual, column=2, value=f"{tanggal}/{bulan}/{tahun}")
    ws.cell(row=row_jual, column=1, value=no)
    for pesan,jumlah in listpesanan.items():
        ws.cell(row=row_jual, column=column, value=jumlah)
        ws.cell(row=row_jual, column=7, value = total)
        column += 1

    print("-----------------------------------------------------")
    print("Pesanan anda :")

    for pesan,jumlah in listpesanan.items():
        if jumlah == 0:
            None
        else:
            print(f"{pesan} : {jumlah}")

    ulangi = input("apakah ada pesanan lagi?")
    if not ulangi:
        continue
    else:
        break
    
wk.save('Data Kasir.xlsx')

# Tutup workbook
wk.close()